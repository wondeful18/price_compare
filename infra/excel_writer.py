from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class ExcelWriter:
    def write(self, frame: pd.DataFrame, target_path: str | Path) -> Path:
        target = Path(target_path)
        target.parent.mkdir(parents=True, exist_ok=True)
        frame.to_excel(target, index=False)
        self._apply_basic_style(target)
        return target

    def _apply_basic_style(self, target: Path) -> None:
        workbook = load_workbook(target)
        sheet = workbook.active
        highlight_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        suspicious_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")

        headers = {cell.value: cell.column for cell in sheet[1]}
        recommended_price_col = headers.get("推荐价格")
        status_col = headers.get("任务状态")
        for row in range(2, sheet.max_row + 1):
            if recommended_price_col and sheet.cell(row=row, column=recommended_price_col).value not in ("", None):
                sheet.cell(row=row, column=recommended_price_col).fill = highlight_fill
            if status_col and sheet.cell(row=row, column=status_col).value == "失败":
                sheet.cell(row=row, column=status_col).fill = suspicious_fill

        sheet.freeze_panes = "A2"
        workbook.save(target)
